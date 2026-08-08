"""
VMB-8A 学习方案B：基于三维材料清单的完整映射
=================================================

核心思路：
  1. 三维材料清单.xlsx提供：零件代号 → U9码映射
  2. 二维材料清单.xls提供：PID图标 → 物料名称 → U9码
  3. BOM点料表.xls提供：U9码 → 物料总数（基准）
  4. 模型库文件名提供：U9码 → .ipt文件映射

学习目标：
  - PID组件类型/管径 → 型号代码映射
  - U9码 → 模型库文件名映射
  - 变体规则（同一U9码如何生成多个实例）
"""

import openpyxl
import json
import os
import re
from collections import defaultdict
from typing import Dict, List, Tuple


def parse_3d_material_list(filepath: str) -> Dict[str, str]:
    """
    解析三维材料清单，提取 零件代号 -> U9码 映射
    
    Args:
        filepath: 三维材料清单.xlsx路径
        
    Returns:
        {零件代号: U9码} 字典
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    part_to_u9 = {}
    for row in range(2, ws.max_row + 1):
        part_name = ws.cell(row, 2).value  # 零件代号
        u9 = ws.cell(row, 5).value  # 库存编号（32）
        
        if part_name and u9:
            part_to_u9[part_name] = str(u9)
    
    return part_to_u9


def parse_2d_material_list(filepath: str) -> Dict[str, Tuple[str, str]]:
    """
    解析二维材料清单，提取 物料名称 -> (U9码, 类型) 映射
    
    表头：序号, 物料名称, 材质, 规格+订货号+描述, 单位, 供应商, U9码（36）, 数量
    索引：  0      1       2           3              4      5        6         7
    
    Args:
        filepath: 二维材料清单.xls路径
        
    Returns:
        {物料名称: (U9码, 类型)} 字典
    """
    import xlrd
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)
    
    name_to_u9 = {}
    for row_idx in range(1, ws.nrows):  # 跳过表头
        name = str(ws.cell_value(row_idx, 1)).strip() if ws.cell_value(row_idx, 1) else ""
        u9 = str(ws.cell_value(row_idx, 6)).strip() if ws.cell_value(row_idx, 6) else ""
        material = str(ws.cell_value(row_idx, 2)).strip() if ws.cell_value(row_idx, 2) else ""  # 材质
        
        if name and u9 and len(u9) >= 12:
            name_to_u9[name] = (u9, material)
    
    return name_to_u9


def parse_model_library(model_dir: str) -> Dict[str, str]:
    """
    解析模型库文件名，提取 U9码 -> 文件名 映射
    
    Args:
        model_dir: 模型库目录路径
        
    Returns:
        {U9码: 文件名} 字典
    """
    u9_to_file = {}
    
    for fname in os.listdir(model_dir):
        if fname.endswith('.ipt') or os.path.isdir(os.path.join(model_dir, fname)):
            # 提取U9码（12位数字）
            match = re.match(r'^(\d{12})', fname)
            if match:
                u9 = match.group(1)
                u9_to_file[u9] = fname
    
    return u9_to_file


def build_complete_mapping(vmb_dir: str) -> dict:
    """
    建立完整的三层映射关系
    
    Args:
        vmb_dir: VMB素材目录路径
        
    Returns:
        完整映射字典
    """
    # 1. 解析三份清单
    bom_path = os.path.join(vmb_dir, "VMB-8A-BOM点料表.xls")
    d2_path = os.path.join(vmb_dir, "VMB-8A-二维图及材料/VMB-8A-二维材料清单.xls")
    d3_path = os.path.join(vmb_dir, "VMB-8A-三维图及材料/VMB-8A-三维材料清单.xlsx")
    model_dir = os.path.join(vmb_dir, "VMB-8A-模型库")
    
    print("解析三份材料清单...")
    d3_mapping = parse_3d_material_list(d3_path)
    d2_mapping = parse_2d_material_list(d2_path)
    model_mapping = parse_model_library(model_dir)
    
    # 2. 建立完整映射
    print("建立完整映射关系...")
    
    # U9码集合（从三份清单收集）
    u9_codes = set(d3_mapping.values()) | set([u9 for u9, _ in d2_mapping.values()])
    
    # 零件代号 -> 模型文件名（通过U9码桥接）
    part_to_model = {}
    for part_name, u9 in d3_mapping.items():
        model_file = model_mapping.get(u9, "")
        if model_file:
            part_to_model[part_name] = model_file
    
    # 3. 统计分析
    print(f"三维材料清单零件数: {len(d3_mapping)}")
    print(f"二维材料清单物料数: {len(d2_mapping)}")
    print(f"模型库文件数: {len(model_mapping)}")
    print(f"唯一U9码数: {len(u9_codes)}")
    print(f"零件→模型映射数: {len(part_to_model)}")
    
    # 4. 导出为知识库格式
    knowledge = {
        "case_name": "VMB-8A",
        "d3_part_to_u9": d3_mapping,  # 零件代号→U9码
        "d2_name_to_u9": d2_mapping,  # 物料名称→U9码
        "u9_to_model": model_mapping,  # U9码→模型文件
        "part_to_model": part_to_model,  # 零件代号→模型文件
        "u9_codes": sorted(list(u9_codes)),  # 唯一U9码列表
        "stats": {
            "d3_parts": len(d3_mapping),
            "d2_materials": len(d2_mapping),
            "models": len(model_mapping),
            "unique_u9": len(u9_codes),
        }
    }
    
    return knowledge


def learn_vmb_8a():
    """学习VMB-8A案例"""
    vmb_dir = r"E:\data\23D转换\VMB素材\VMB素材"
    
    print("="*60)
    print("学习 VMB-8A 案例（方案B：基于三维材料清单）")
    print("="*60)
    
    knowledge = build_complete_mapping(vmb_dir)
    
    # 保存知识库
    output_path = r"E:\workbuddy\2026-07-31-15-46-13\23d_converter\data\knowledge\VMB-8A_u9_mapping.json"
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(knowledge, f, ensure_ascii=False, indent=2)
    
    print(f"\n知识库已保存: {output_path}")
    print("\n下一步：用此映射关系更新学习引擎")
    
    return knowledge


if __name__ == "__main__":
    knowledge = learn_vmb_8a()